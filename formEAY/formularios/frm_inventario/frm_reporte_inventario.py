# -*- coding: utf-8 -*-

###########################################################################
## Python code generated with wxFormBuilder (version Jun 17 2015)
## http://www.wxformbuilder.org/
##
## PLEASE DO "NOT" EDIT THIS FILE!
###########################################################################

import wx
import wx.xrc
import wx.grid
import os
import pandas as pd

from pyeay.grillas import ManipularGrillas
from pyeay.dbcac.conexiondb import Ejecutar_SQL
from formEAY.constantesCAC.constantesCAC import BasesDeDatos

from formEAY.constantesCAC.imgCAC import Img_formularios_general


###########################################################################
## Class ReporteInventario
###########################################################################

class ReporteInventario(wx.Frame):

    def __init__(self, parent):
        wx.Frame.__init__(self, parent, id=wx.ID_ANY, title=u"Reporte de Inventario", pos=wx.DefaultPosition,
                          size=wx.Size(762, 729), style=wx.DEFAULT_FRAME_STYLE | wx.TAB_TRAVERSAL)

        self.SetSizeHints(wx.Size(600, 400), wx.DefaultSize)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))

        icono_formularios = Img_formularios_general()

        self.orden_ascendente = True
        self.columna = 0

        bSizer23 = wx.BoxSizer(wx.VERTICAL)

        self.m_panel_principal = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL)
        self.m_panel_principal.SetBackgroundColour(wx.Colour(255, 255, 255))

        bSizer_principal = wx.BoxSizer(wx.VERTICAL)

        bSizer25 = wx.BoxSizer(wx.HORIZONTAL)

        self.lbl_titulo = wx.StaticText(self.m_panel_principal, wx.ID_ANY, u"Reporte de Inventario", wx.DefaultPosition,
                                        wx.DefaultSize, wx.ALIGN_CENTRE)
        self.lbl_titulo.Wrap(-1)
        self.lbl_titulo.SetFont(wx.Font(11, 70, 90, 92, False, wx.EmptyString))
        self.lbl_titulo.SetForegroundColour(wx.Colour(255, 128, 64))

        bSizer25.Add(self.lbl_titulo, 1, wx.ALL, 5)

        bSizer26 = wx.BoxSizer(wx.HORIZONTAL)

        self.btn_exportar_xlsx = wx.Button(self.m_panel_principal, wx.ID_ANY, u"--> &XLSX", wx.DefaultPosition,
                                           wx.DefaultSize, wx.NO_BORDER)
        self.btn_exportar_xlsx.SetBackgroundColour(wx.Colour(0, 255, 0))

        bSizer26.Add(self.btn_exportar_xlsx, 0, wx.ALIGN_RIGHT | wx.ALL, 5)

        self.btn_actualizar = wx.BitmapButton(self.m_panel_principal, wx.ID_ANY, wx.Bitmap(
            icono_formularios.ACTUALIZAR_FORMULARIO, wx.BITMAP_TYPE_ANY),
                                              wx.DefaultPosition, wx.DefaultSize, wx.BU_AUTODRAW | wx.NO_BORDER)

        self.btn_actualizar.SetBitmapCurrent(
            wx.Bitmap(icono_formularios.ACTUALIZAR_FORMULARIO_SEL, wx.BITMAP_TYPE_ANY))
        bSizer26.Add(self.btn_actualizar, 0, wx.ALL, 5)

        bSizer25.Add(bSizer26, 0, 0, 5)

        bSizer_principal.Add(bSizer25, 0, wx.EXPAND, 5)

        bSizer28 = wx.BoxSizer(wx.VERTICAL)

        self.m_staticline1 = wx.StaticLine(self.m_panel_principal, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                           wx.LI_HORIZONTAL)
        bSizer28.Add(self.m_staticline1, 0, wx.EXPAND | wx.ALL, 5)

        bSizer_principal.Add(bSizer28, 0, wx.EXPAND, 5)

        self.lbl_etq_totales = wx.StaticText(self.m_panel_principal, wx.ID_ANY, u"Totales", wx.DefaultPosition,
                                             wx.DefaultSize, 0)
        self.lbl_etq_totales.Wrap(-1)
        bSizer_principal.Add(self.lbl_etq_totales, 0, wx.ALL, 5)

        self.grid_resumen = wx.grid.Grid(self.m_panel_principal, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, 0)

        # Grid
        self.grid_resumen.CreateGrid(1, 7)
        self.grid_resumen.EnableEditing(False)
        self.grid_resumen.EnableGridLines(True)
        self.grid_resumen.EnableDragGridSize(False)
        self.grid_resumen.SetMargins(0, 0)

        # Columns
        self.grid_resumen.AutoSizeColumns()
        self.grid_resumen.EnableDragColMove(False)
        self.grid_resumen.EnableDragColSize(True)
        self.grid_resumen.SetColLabelSize(30)
        self.grid_resumen.SetColLabelValue(0, u"Unid 1ra")
        self.grid_resumen.SetColLabelValue(1, u"Unid 2da")
        self.grid_resumen.SetColLabelValue(2, u"Unid Extusion")
        self.grid_resumen.SetColLabelValue(3, u"Unid Cargue")
        self.grid_resumen.SetColLabelValue(4, u"Ton 1ra")
        self.grid_resumen.SetColLabelValue(5, u"Ton 2da")
        self.grid_resumen.SetColLabelValue(6, u"Total Ton")
        self.grid_resumen.SetColLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Rows
        self.grid_resumen.EnableDragRowSize(True)
        self.grid_resumen.SetRowLabelSize(50)
        self.grid_resumen.SetRowLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Label Appearance
        self.grid_resumen.SetLabelBackgroundColour(wx.Colour(255, 255, 255))

        # Cell Defaults
        self.grid_resumen.SetDefaultCellAlignment(wx.ALIGN_LEFT, wx.ALIGN_TOP)
        self.grid_resumen.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_INFOBK))

        bSizer_principal.Add(self.grid_resumen, 0, wx.ALL, 5)

        self.lbl_etq_detalle = wx.StaticText(self.m_panel_principal, wx.ID_ANY, u"Detalle", wx.DefaultPosition,
                                             wx.DefaultSize, 0)
        self.lbl_etq_detalle.Wrap(-1)
        bSizer_principal.Add(self.lbl_etq_detalle, 0, wx.ALL, 5)

        self.grid_productos = wx.grid.Grid(self.m_panel_principal, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, 0)

        # Grid
        self.grid_productos.CreateGrid(0, 10)
        self.grid_productos.EnableEditing(False)
        self.grid_productos.EnableGridLines(True)
        self.grid_productos.EnableDragGridSize(False)
        self.grid_productos.SetMargins(0, 0)

        # Columns
        self.grid_productos.AutoSizeColumns()
        self.grid_productos.EnableDragColMove(True)
        self.grid_productos.EnableDragColSize(True)
        self.grid_productos.SetColLabelSize(30)
        self.grid_productos.SetColLabelValue(0, u"id")
        self.grid_productos.SetColLabelValue(1, u"Producto")
        self.grid_productos.SetColLabelValue(2, u"Cat")
        self.grid_productos.SetColLabelValue(3, u"Unid 1ra")
        self.grid_productos.SetColLabelValue(4, u"Unid 2da")
        self.grid_productos.SetColLabelValue(5, u"Unid Extusion")
        self.grid_productos.SetColLabelValue(6, u"Unid Cargue")
        self.grid_productos.SetColLabelValue(7, u"Ton 1ra")
        self.grid_productos.SetColLabelValue(8, u"Ton 2da")
        self.grid_productos.SetColLabelValue(9, u"Total Ton")
        self.grid_productos.SetColLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Rows
        self.grid_productos.EnableDragRowSize(True)
        self.grid_productos.SetRowLabelSize(35)
        self.grid_productos.SetRowLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Label Appearance
        self.grid_productos.SetLabelBackgroundColour(wx.Colour(255, 255, 255))

        # Cell Defaults
        self.grid_productos.SetDefaultCellAlignment(wx.ALIGN_LEFT, wx.ALIGN_TOP)
        self.grid_productos.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_INFOBK))

        bSizer_principal.Add(self.grid_productos, 1, wx.ALL | wx.EXPAND, 5)

        self.m_panel_principal.SetSizer(bSizer_principal)
        self.m_panel_principal.Layout()
        bSizer_principal.Fit(self.m_panel_principal)
        bSizer23.Add(self.m_panel_principal, 1, wx.EXPAND | wx.ALL, 5)

        self.SetSizer(bSizer23)

        ## VALORTES INICIALES EAY
        self.func_valores_iniciales()


        self.Layout()

        self.Centre(wx.BOTH)

        # Connect Events
        self.btn_exportar_xlsx.Bind(wx.EVT_BUTTON, self.btn_exportar_xlsxOnButtonClick)
        self.btn_actualizar.Bind(wx.EVT_BUTTON, self.btn_actualizarOnButtonClick)
        self.grid_resumen.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.grid_productosOnGridCellLeftClick)
        self.grid_productos.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.grid_productosOnGridCellLeftClick)

        self.grid_productos.Bind(wx.grid.EVT_GRID_LABEL_LEFT_DCLICK, self.grid_productosOnGridLabelLeftDClick)

    def __del__(self):
        pass

    # Virtual event handlers, overide them in your derived class

    def grid_productosOnGridLabelLeftDClick(self, event):
        columna = event.GetCol()

        if self.orden_ascendente == True:
            self.orden_ascendente = False
        else:
            self.orden_ascendente = True

        self.func_ordenarGrillaPorColumna(columna, self.orden_ascendente)
        event.Skip()

    def btn_exportar_xlsxOnButtonClick(self, event):
        nom_archivo, la_ruta = self.func_crear_xlsx()

        ruta_completa  = os.path.join(la_ruta, nom_archivo)

        os.startfile(ruta_completa)

        event.Skip()

    def btn_actualizarOnButtonClick(self, event):
        self.func_valores_iniciales()
        event.Skip()

    def grid_productosOnGridCellLeftClick(self, event):
        event.Skip()

    ## FUNCIONES EAY



    def func_crear_xlsx(self):
        from openpyxl import Workbook
        from openpyxl.styles import colors
        from openpyxl.styles import Font, Color
        from openpyxl.styles import PatternFill
        import os
        from datetime import datetime
        import shutil

        wb = Workbook()
        # grab the active worksheet
        ws = wb.active

        relleno = PatternFill(start_color='F2DDDC', end_color='F2DDDC', fill_type='solid')  # 'FFAFEEEE'

        ft_titulo = Font(color="BF0041", size=14, bold=True)
        ft_cabecera = Font(color="BF0041", size=12, bold=True)

        ws.title = "Reporte General de Inventario"

        cabeceras = ['id', 'Producto', 'Categoria', 'Unid 1ra', 'Unid 2da', 'Unid Extusion', 'Unid Cargue',
                      'Ton 1ra', 'Ton 2da', 'Total Ton']
        ##  --------------------------------

        nomCarpeta = 'REPORTES\XLSX\INVENTARIO'
        nomArchivo = 'Reporte General de Inventario'

        try:
            dir_trabajo = os.getcwd()
            rutaCarpeta = os.path.join(dir_trabajo, nomCarpeta)

            if not os.path.exists(nomCarpeta):
                os.mkdir(nomCarpeta)
            la_ruta = os.path.join(dir_trabajo, nomCarpeta)

            ahora = datetime.now()

            la_fecha = str(ahora.year) + '-' + str(ahora.month) + '-' + str(ahora.day) + '  HH--' + str(
                ahora.hour) + '-' + str(ahora.minute) + '-' + str(ahora.second)

            nombreCompletoArchivo = nomArchivo + ' ' + la_fecha + u'.xlsx'

            # Data can be assigned directly to cells
            # sheet['A1'] = 42
            ws['A1'] = nomArchivo
            ws['A2'] = 'Fecha:'
            ws['A3'] = 'Hora:'
            ws['B2'] = ahora.date()
            ws['B3'] = ahora.time()

            ws['A4'] = ''  # filas vacias pora dar orden

            cabeceras_resumen = ['Unid 1ra', 'Unid 2da', 'Unid Extusion', 'Unid Cargue', 'Ton 1ra', 'Ton 2da', 'Total Ton']

            ws.append(cabeceras_resumen)
            ws.append(self.rows_resumen[0])

            ws['A7'] = ''

            titulo = ws['A1']
            titulo.font = ft_titulo

            ws.merge_cells('A1:O1')

            ws.merge_cells('B2:O2')
            ws.merge_cells('B3:O3')

            ws.merge_cells('A4:I4')

            for j in range(len(cabeceras_resumen)):
                cell_range = ws.cell(row=5, column=j + 1)  # era row=1
                cell_range.fill = relleno

            ws.append(cabeceras)
            rows_datos = self.rows

            for fila in rows_datos:
                ws.append(fila)

            for j in range(len(cabeceras)):
                cell_range = ws.cell(row=8, column=j + 1)  # era row=1
                cell_range.fill = relleno

            cell_range = ws.cell(row=1, column=1)  # era row=1
            cell_range.fill = relleno

            from openpyxl.styles import Alignment
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['B2'].alignment = Alignment(horizontal="left")
            ws['B3'].alignment = Alignment(horizontal="left")

            wb.save(nombreCompletoArchivo)

            shutil.move(nombreCompletoArchivo, la_ruta)

            wx.MessageBox(u'Se ha realizado con exito la Exportación', u'Atención', wx.OK | wx.ICON_INFORMATION)

            return nombreCompletoArchivo, la_ruta
        except:
            return 0

    def func_valores_iniciales(self):
        self.func_cargar_grilla_productos()
        self.grid_productos.AutoSizeColumns()
        self.grid_resumen.AutoSizeColumns()

        # list_columnas  = [7, 8, 9, 10, 11]
        # ManipularGrillas.ocultarColumnasGrilla(self.grid_productos, list_columnas)
        #
        # list_columnas = [4, 5, 6]
        # ManipularGrillas.ocultarColumnasGrilla(self.grid_resumen, list_columnas)

    def func_cargar_grilla_productos(self):

        sSql = """
                SELECT id_producto, nom_producto, categoria, stock_primera, stock_segunda, stock_extrusion, stock_cargue,                        
                        TRUNC(peso * stock_primera / 1000000.0, 2) as ton_1ra, 
                        TRUNC(peso * stock_segunda / 1000000.0, 2) as ton_2da, 
                        TRUNC(((peso * stock_primera)+(peso * stock_segunda)) / 1000000.0, 2) as total_ton
                FROM producto
                WHERE activo = 'True'
        """

        cabeceras = ['id', 'Producto', 'Categoria', 'Unid 1ra', 'Unid 2da', 'Unid Extusion', 'Unid Cargue', 'Ton 1ra', 'Ton 2da', 'Total Ton']


        self.rows = Ejecutar_SQL.select_varios_registros(sSql, 'frm_reporte_inventario/func_cargar_grilla_productos', 500, BasesDeDatos.DB_PRINCIPAL)


        df = pd.DataFrame(self.rows, columns = cabeceras)

        #df = df.groupby(['id']).sum()
        df = df[['Unid 1ra', 'Unid 2da', 'Unid Extusion', 'Unid Cargue', 'Ton 1ra', 'Ton 2da', 'Total Ton']].sum()

        self.rows_resumen = []
        self.rows_resumen.append(df.tolist())

        ManipularGrillas.llenarGrilla(self.grid_productos, self.rows)
        ManipularGrillas.llenarGrilla(self.grid_resumen, self.rows_resumen)

        self.func_colorear_grilla()


    def func_ordenarGrillaPorColumna(self, columna, orden_ascendente):

        self.columna = columna

        lista_tipoColumna = ['int', 'str', 'str', 'int', 'int', 'int', 'int',
                             'float', 'float', 'float']

        ManipularGrillas.ordenarGrillaPorColumna(self.grid_productos,  self.columna, lista_tipoColumna,
                                                 orden_ascendente)

        self.func_colorear_grilla()

    def func_colorear_grilla(self):
        UNID_PRIMERA_AZUL = wx.Colour(255, 213, 177)
        UNID_SEGUNDA_AZUL = wx.Colour(255, 236, 234)

        # PVP_PRIMERA_VERDE = wx.Colour(255, 236, 234)
        # PVP_SEGUNDA_VERDE = wx.Colour(229, 212, 210)
        # PVP_TOTAL_VERDE = wx.Colour(204, 188, 187)

        TON_PRIMERA_LILA = wx.Colour(193, 152, 177)
        TON_SEGUNDA_LILA = wx.Colour(202, 168, 189)
        TON_TOTAL_LILA = wx.Colour(220, 198, 211)


        # dic_color = {3: UNID_PRIMERA_AZUL, 4: UNID_SEGUNDA_AZUL, 9:PVP_PRIMERA_VERDE, 10:PVP_SEGUNDA_VERDE, 11:PVP_TOTAL_VERDE,
        #              12: TON_PRIMERA_LILA, 13: TON_SEGUNDA_LILA, 14: TON_TOTAL_LILA
        #              }

        dic_color = {3: UNID_PRIMERA_AZUL, 4: UNID_SEGUNDA_AZUL,
                     7: TON_PRIMERA_LILA, 8: TON_SEGUNDA_LILA, 9: TON_TOTAL_LILA }
        ManipularGrillas.setColorFondoCeldaGrilla(self.grid_productos, dic_color)



