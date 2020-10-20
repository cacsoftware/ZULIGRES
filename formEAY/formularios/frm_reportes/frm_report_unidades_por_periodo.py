# -*- coding: utf-8 -*-

###########################################################################
## Python code generated with wxFormBuilder (version Jun 17 2015)
## http://www.wxformbuilder.org/
##
## PLEASE DO "NOT" EDIT THIS FILE!
###########################################################################

import wx
import wx.xrc
import wx.grid
import wx.adv
import numpy as np

from pyeay.rows import ManipularRows
from pyeay.grillas import ManipularGrillas
from pyeay.dbcac.conexiondb import Ejecutar_SQL

from formEAY.constantesCAC.constantesCAC import BasesDeDatos
from formEAY.dbaseCAC.dbVarios import DbGetVarios
from formEAY.constantesCAC.constantesCAC import AreasProduccion
from formEAY.constantesCAC.coloresCAC import ColorsFondoCellGrilla

COLOR_RESALTE_PRIMERA = wx.Colour(255, 236, 234)
COLOR_RESALTE_SEGUNDA = wx.Colour(229, 212, 210)

COLOR_RESALTE_ROTURA = wx.Colour(233, 192, 192)  ## ROTURA
COLOR_RESALTE_ROTURA_CLARO = ColorsFondoCellGrilla.RESALTE_ROSA2

COLOR_RESALTE_TOTAL = wx.Colour(255, 213, 177)

AREA_EXTRUSION = AreasProduccion.EXTRUSION
AREA_CARGUE_VAGONETAS = AreasProduccion.CARGUE_VAGONETAS
AREA_DESCARGUE_VAGONETAS = AreasProduccion.DESCARGUE_VAGONETAS

###########################################################################
## Class UnidadesPorPeriodo
###########################################################################

class UnidadesPorPeriodo(wx.Frame):

    def __init__(self, parent):
        wx.Frame.__init__(self, parent, id=wx.ID_ANY,
                          title=u"Producción por Periodo [Extrusión, Cargue y descargue de Vagonetas]",
                          pos=wx.DefaultPosition, size=wx.Size(1150, 650),
                          style=wx.DEFAULT_FRAME_STYLE | wx.TAB_TRAVERSAL)

        self.SetSizeHints(wx.Size(950, 400), wx.DefaultSize)
        self.SetBackgroundColour(wx.Colour(240, 240, 240))

        self.orden_ascendente = True

        bSizer_principal = wx.BoxSizer(wx.VERTICAL)

        bSizer_panelCabecera = wx.BoxSizer(wx.HORIZONTAL)

        self.panel_cabecera = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL)
        self.panel_cabecera.SetBackgroundColour(wx.Colour(255, 255, 255))

        bSizer_cabecera = wx.BoxSizer(wx.HORIZONTAL)

        bSizer_rango_fechas = wx.BoxSizer(wx.VERTICAL)

        self.lbl_etq_rangoFechas = wx.StaticText(self.panel_cabecera, wx.ID_ANY, u"Rango Fechas", wx.DefaultPosition,
                                                 wx.DefaultSize, 0)
        self.lbl_etq_rangoFechas.Wrap(-1)
        self.lbl_etq_rangoFechas.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_etq_rangoFechas.SetForegroundColour(wx.Colour(255, 128, 0))

        bSizer_rango_fechas.Add(self.lbl_etq_rangoFechas, 0, wx.ALL, 5)

        bSizer_fecha1 = wx.BoxSizer(wx.HORIZONTAL)

        self.lbl_etq_fechaInicial = wx.StaticText(self.panel_cabecera, wx.ID_ANY, u"Fecha Inicial:", wx.DefaultPosition,
                                                  wx.Size(70, -1), wx.ALIGN_RIGHT)
        self.lbl_etq_fechaInicial.Wrap(-1)
        bSizer_fecha1.Add(self.lbl_etq_fechaInicial, 0, wx.ALL, 5)

        self.datePicker_fecha1 = wx.adv.DatePickerCtrl(self.panel_cabecera, wx.ID_ANY, wx.DefaultDateTime,
                                                   wx.DefaultPosition, wx.DefaultSize, style = wx.adv.DP_DROPDOWN | wx.adv.DP_SHOWCENTURY)
        bSizer_fecha1.Add(self.datePicker_fecha1, 1, wx.ALL, 5)

        bSizer_rango_fechas.Add(bSizer_fecha1, 0, wx.EXPAND, 5)

        bSizer_fecha2 = wx.BoxSizer(wx.HORIZONTAL)

        self.lbl_etq_fechaFinal = wx.StaticText(self.panel_cabecera, wx.ID_ANY, u"Fecha Final:", wx.DefaultPosition,
                                                wx.Size(70, -1), wx.ALIGN_RIGHT)
        self.lbl_etq_fechaFinal.Wrap(-1)
        bSizer_fecha2.Add(self.lbl_etq_fechaFinal, 0, wx.ALL, 5)

        self.datePicker_fecha2 = wx.adv.DatePickerCtrl(self.panel_cabecera, wx.ID_ANY, wx.DefaultDateTime,
                                                   wx.DefaultPosition, wx.DefaultSize, style = wx.adv.DP_DROPDOWN | wx.adv.DP_SHOWCENTURY)
        bSizer_fecha2.Add(self.datePicker_fecha2, 1, wx.ALL, 5)

        bSizer_rango_fechas.Add(bSizer_fecha2, 1, wx.EXPAND, 5)

        bSizer_cabecera.Add(bSizer_rango_fechas, 1, wx.EXPAND, 5)

        bSizer_area = wx.BoxSizer(wx.VERTICAL)

        self.lbl_etq_area = wx.StaticText(self.panel_cabecera, wx.ID_ANY, u"Area", wx.DefaultPosition, wx.DefaultSize,
                                          0)
        self.lbl_etq_area.Wrap(-1)
        self.lbl_etq_area.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_etq_area.SetForegroundColour(wx.Colour(255, 128, 0))

        bSizer_area.Add(self.lbl_etq_area, 0, wx.ALL, 5)

        bSizer17 = wx.BoxSizer(wx.VERTICAL)

        radioBox_areaChoices = [u"EXTRUSION", u"CARGUE DE VAGONETAS",
                                u"DESCARGUE DE VAGONETAS", u"DESPACHOS"]
        self.radioBox_area = wx.RadioBox(self.panel_cabecera, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                         wx.DefaultSize, radioBox_areaChoices, 1, wx.RA_SPECIFY_COLS)
        self.radioBox_area.SetSelection(0)
        bSizer17.Add(self.radioBox_area, 0, wx.ALL, 5)

        bSizer_area.Add(bSizer17, 1, wx.EXPAND, 5)

        bSizer_cabecera.Add(bSizer_area, 0, wx.EXPAND, 5)

        bSizer_estado = wx.BoxSizer(wx.VERTICAL)

        self.lbl_etq_periodo = wx.StaticText(self.panel_cabecera, wx.ID_ANY, u"Periodo", wx.DefaultPosition,
                                             wx.DefaultSize, 0)
        self.lbl_etq_periodo.Wrap(-1)
        self.lbl_etq_periodo.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_etq_periodo.SetForegroundColour(wx.Colour(255, 128, 0))

        bSizer_estado.Add(self.lbl_etq_periodo, 0, wx.ALL, 5)

        radioBox_periodoChoices = [u"DIA", u"SEMANA", u"MES", u"AÑO"]
        self.radioBox_periodo = wx.RadioBox(self.panel_cabecera, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                            wx.DefaultSize, radioBox_periodoChoices, 1, wx.RA_SPECIFY_COLS)
        self.radioBox_periodo.SetSelection(1)
        bSizer_estado.Add(self.radioBox_periodo, 0, wx.ALL, 5)

        bSizer_cabecera.Add(bSizer_estado, 0, wx.EXPAND, 5)

        bSizer_turno = wx.BoxSizer(wx.VERTICAL)

        self.lbl_etq_turno = wx.StaticText(self.panel_cabecera, wx.ID_ANY, u"Turno", wx.DefaultPosition, wx.DefaultSize,
                                           0)
        self.lbl_etq_turno.Wrap(-1)
        self.lbl_etq_turno.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_etq_turno.SetForegroundColour(wx.Colour(255, 128, 0))

        bSizer_turno.Add(self.lbl_etq_turno, 0, wx.ALL, 5)

        checkList_turnoChoices = [u"turno 1", u"turno 2"]
        self.checkList_turno = wx.CheckListBox(self.panel_cabecera, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                               checkList_turnoChoices, 0)
        bSizer_turno.Add(self.checkList_turno, 1, wx.ALL | wx.EXPAND, 5)

        bSizer_cabecera.Add(bSizer_turno, 2, wx.EXPAND, 5)

        bSizer_botonesCabecera = wx.BoxSizer(wx.VERTICAL)

        self.m_staticText6 = wx.StaticText(self.panel_cabecera, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                           wx.DefaultSize, 0)
        self.m_staticText6.Wrap(-1)
        bSizer_botonesCabecera.Add(self.m_staticText6, 0, wx.ALL, 5)

        self.btn_buscar = wx.Button(self.panel_cabecera, wx.ID_ANY, u"&Buscar", wx.DefaultPosition, wx.DefaultSize,
                                    wx.NO_BORDER)
        self.btn_buscar.SetBackgroundColour(wx.Colour(0, 255, 255))

        bSizer_botonesCabecera.Add(self.btn_buscar, 0, wx.ALL, 5)

        self.btn_graficar = wx.Button(self.panel_cabecera, wx.ID_ANY, u"&Graficar", wx.DefaultPosition, wx.DefaultSize,
                                      wx.NO_BORDER)
        self.btn_graficar.SetForegroundColour(wx.Colour(255, 255, 255))
        self.btn_graficar.SetBackgroundColour(wx.Colour(255, 128, 0))

        bSizer_botonesCabecera.Add(self.btn_graficar, 0, wx.ALL, 5)

        bSizer_cabecera.Add(bSizer_botonesCabecera, 0, wx.EXPAND, 5)

        self.panel_cabecera.SetSizer(bSizer_cabecera)
        self.panel_cabecera.Layout()
        bSizer_cabecera.Fit(self.panel_cabecera)
        bSizer_panelCabecera.Add(self.panel_cabecera, 1, wx.EXPAND | wx.ALL, 5)

        bSizer_principal.Add(bSizer_panelCabecera, 0, wx.EXPAND, 5)

        bSizer_panelResultados = wx.BoxSizer(wx.VERTICAL)

        self.m_panel_resultados = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL)
        self.m_panel_resultados.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))

        bSizer_resultados = wx.BoxSizer(wx.VERTICAL)

        bSizer_totalesConsulta = wx.BoxSizer(wx.HORIZONTAL)

        self.lbl_etq_area_resultados = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"Area:", wx.DefaultPosition,
                                                     wx.DefaultSize, 0)
        self.lbl_etq_area_resultados.Wrap(-1)
        bSizer_totalesConsulta.Add(self.lbl_etq_area_resultados, 0, wx.ALL, 5)

        self.lbl_area = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"DESCARGUE DE VAGONETAS", wx.DefaultPosition,
                                      wx.DefaultSize, 0)
        self.lbl_area.Wrap(-1)
        self.lbl_area.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_area.SetForegroundColour(wx.Colour(0, 0, 0))
        self.lbl_area.SetMinSize(wx.Size(170, -1))

        bSizer_totalesConsulta.Add(self.lbl_area, 0, wx.ALL, 5)

        self.lbl_etq_fechas = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"Fechas:", wx.DefaultPosition,
                                            wx.DefaultSize, 0)
        self.lbl_etq_fechas.Wrap(-1)
        bSizer_totalesConsulta.Add(self.lbl_etq_fechas, 0, wx.ALL, 5)

        self.lbl_fechas = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"25/10/2020  al 31/10/2020",
                                        wx.DefaultPosition, wx.DefaultSize, 0)
        self.lbl_fechas.Wrap(-1)
        self.lbl_fechas.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_fechas.SetMinSize(wx.Size(150, -1))

        bSizer_totalesConsulta.Add(self.lbl_fechas, 0, wx.ALL, 5)

        self.lbl_etq_periodo = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"Periodo:", wx.DefaultPosition,
                                             wx.DefaultSize, 0)
        self.lbl_etq_periodo.Wrap(-1)
        bSizer_totalesConsulta.Add(self.lbl_etq_periodo, 0, wx.ALL, 5)

        self.lbl_periodo = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"SEMANA", wx.DefaultPosition,
                                         wx.DefaultSize, 0)
        self.lbl_periodo.Wrap(-1)
        self.lbl_periodo.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_periodo.SetMinSize(wx.Size(60, -1))

        bSizer_totalesConsulta.Add(self.lbl_periodo, 0, wx.ALL, 5)

        self.lbl_etq_turnos = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"Turnos:", wx.DefaultPosition,
                                            wx.Size(50, -1), wx.ALIGN_RIGHT)
        self.lbl_etq_turnos.Wrap(-1)
        bSizer_totalesConsulta.Add(self.lbl_etq_turnos, 0, wx.ALL, 5)

        self.lbl_turnos = wx.StaticText(self.m_panel_resultados, wx.ID_ANY, u"Am", wx.DefaultPosition, wx.DefaultSize,
                                        0)
        self.lbl_turnos.Wrap(-1)
        self.lbl_turnos.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        self.lbl_turnos.SetMinSize(wx.Size(60, -1))

        bSizer_totalesConsulta.Add(self.lbl_turnos, 1, wx.ALL, 5)

        bSizer_resultados.Add(bSizer_totalesConsulta, 0, wx.EXPAND, 5)

        bSizer15 = wx.BoxSizer(wx.HORIZONTAL)

        bSizer172 = wx.BoxSizer(wx.VERTICAL)

        self.m_panel3 = wx.Panel(self.m_panel_resultados, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                 wx.TAB_TRAVERSAL)
        self.m_panel3.SetBackgroundColour(wx.Colour(240, 240, 240))

        bSizer18 = wx.BoxSizer(wx.VERTICAL)

        self.grid_resultado_busqueda = wx.grid.Grid(self.m_panel3, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, 0)

        # Grid
        self.grid_resultado_busqueda.CreateGrid(0, 4)
        self.grid_resultado_busqueda.EnableEditing(False)
        self.grid_resultado_busqueda.EnableGridLines(True)
        self.grid_resultado_busqueda.EnableDragGridSize(False)
        self.grid_resultado_busqueda.SetMargins(0, 0)

        # Columns
        self.grid_resultado_busqueda.AutoSizeColumns()
        self.grid_resultado_busqueda.EnableDragColMove(False)
        self.grid_resultado_busqueda.EnableDragColSize(True)
        self.grid_resultado_busqueda.SetColLabelSize(30)
        self.grid_resultado_busqueda.SetColLabelValue(0, u"Semana")
        self.grid_resultado_busqueda.SetColLabelValue(1, u"Primera")
        self.grid_resultado_busqueda.SetColLabelValue(2, u"Segunda")
        self.grid_resultado_busqueda.SetColLabelValue(3, u"Rotura")
        self.grid_resultado_busqueda.SetColLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Rows
        self.grid_resultado_busqueda.EnableDragRowSize(True)
        self.grid_resultado_busqueda.SetRowLabelSize(50)
        self.grid_resultado_busqueda.SetRowLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Label Appearance
        self.grid_resultado_busqueda.SetLabelBackgroundColour(wx.Colour(255, 255, 255))

        # Cell Defaults
        self.grid_resultado_busqueda.SetDefaultCellBackgroundColour(wx.Colour(255, 255, 255))
        self.grid_resultado_busqueda.SetDefaultCellAlignment(wx.ALIGN_LEFT, wx.ALIGN_TOP)
        bSizer18.Add(self.grid_resultado_busqueda, 1, wx.ALL | wx.EXPAND, 5)

        self.m_panel3.SetSizer(bSizer18)
        self.m_panel3.Layout()
        bSizer18.Fit(self.m_panel3)
        bSizer172.Add(self.m_panel3, 1, wx.EXPAND | wx.ALL, 5)

        bSizer15.Add(bSizer172, 3, wx.EXPAND, 5)

        bSizer19 = wx.BoxSizer(wx.VERTICAL)

        self.grid_estadisticas = wx.grid.Grid(self.m_panel_resultados, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, 0)

        # Grid
        self.grid_estadisticas.CreateGrid(7, 4)
        self.grid_estadisticas.EnableEditing(False)
        self.grid_estadisticas.EnableGridLines(True)
        self.grid_estadisticas.EnableDragGridSize(False)
        self.grid_estadisticas.SetMargins(0, 0)

        # Columns
        self.grid_estadisticas.AutoSizeColumns()
        self.grid_estadisticas.EnableDragColMove(False)
        self.grid_estadisticas.EnableDragColSize(True)
        self.grid_estadisticas.SetColLabelSize(30)
        self.grid_estadisticas.SetColLabelValue(0, u"Total Unid")
        self.grid_estadisticas.SetColLabelValue(1, u"Primera")
        self.grid_estadisticas.SetColLabelValue(2, u"Segunda")
        self.grid_estadisticas.SetColLabelValue(3, u"Rotura")
        self.grid_estadisticas.SetColLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Rows
        self.grid_estadisticas.EnableDragRowSize(True)
        self.grid_estadisticas.SetRowLabelSize(90)
        self.grid_estadisticas.SetRowLabelValue(0, u"Sumatoria")
        self.grid_estadisticas.SetRowLabelValue(1, u"Media")
        self.grid_estadisticas.SetRowLabelValue(2, u"Desv. Estandar")
        self.grid_estadisticas.SetRowLabelValue(3, u"Mediana")
        self.grid_estadisticas.SetRowLabelValue(4, u"Máximo")
        self.grid_estadisticas.SetRowLabelValue(5, u"Mínimo")
        self.grid_estadisticas.SetRowLabelValue(6, u"n")
        self.grid_estadisticas.SetRowLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Label Appearance
        self.grid_estadisticas.SetLabelBackgroundColour(wx.Colour(255, 255, 255))

        # Cell Defaults
        self.grid_estadisticas.SetDefaultCellBackgroundColour(wx.Colour(255, 255, 255))
        self.grid_estadisticas.SetDefaultCellAlignment(wx.ALIGN_LEFT, wx.ALIGN_TOP)
        bSizer19.Add(self.grid_estadisticas, 0, wx.ALL|wx.EXPAND, 5 )

        #---------------------------
        self.grid_estadisticas2 = wx.grid.Grid(self.m_panel_resultados, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                               0)

        # Grid
        self.grid_estadisticas2.CreateGrid(7, 4)
        self.grid_estadisticas2.EnableEditing(False)
        self.grid_estadisticas2.EnableGridLines(True)
        self.grid_estadisticas2.EnableDragGridSize(False)
        self.grid_estadisticas2.SetMargins(0, 0)

        # Columns
        self.grid_estadisticas2.AutoSizeColumns()
        self.grid_estadisticas2.EnableDragColMove(False)
        self.grid_estadisticas2.EnableDragColSize(True)
        self.grid_estadisticas2.SetColLabelSize(30)
        self.grid_estadisticas2.SetColLabelValue(0, u"Total Ton ")
        self.grid_estadisticas2.SetColLabelValue(1, u"Ton 1ra ")
        self.grid_estadisticas2.SetColLabelValue(2, u"Ton 2da ")
        self.grid_estadisticas2.SetColLabelValue(3, u"Ton Rotura ")
        self.grid_estadisticas2.SetColLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Rows
        self.grid_estadisticas2.EnableDragRowSize(True)
        self.grid_estadisticas2.SetRowLabelSize(90)
        self.grid_estadisticas2.SetRowLabelValue(0, u"Sumatoria")
        self.grid_estadisticas2.SetRowLabelValue(1, u"Media")
        self.grid_estadisticas2.SetRowLabelValue(2, u"Desv. Estandar")
        self.grid_estadisticas2.SetRowLabelValue(3, u"Mediana")
        self.grid_estadisticas2.SetRowLabelValue(4, u"Máximo")
        self.grid_estadisticas2.SetRowLabelValue(5, u"Mínimo")
        self.grid_estadisticas2.SetRowLabelValue(6, u"n")
        self.grid_estadisticas2.SetRowLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Label Appearance
        self.grid_estadisticas2.SetLabelBackgroundColour(wx.Colour(255, 255, 255))

        # Cell Defaults
        self.grid_estadisticas2.SetDefaultCellBackgroundColour(wx.Colour(255, 255, 255))
        self.grid_estadisticas2.SetDefaultCellAlignment(wx.ALIGN_LEFT, wx.ALIGN_TOP)
        bSizer19.Add(self.grid_estadisticas2, 1, wx.ALL, 5)

        #---------------------------

        bSizer15.Add(bSizer19, 2, wx.EXPAND, 5)

        bSizer_resultados.Add(bSizer15, 1, wx.EXPAND, 5)

        self.m_panel_resultados.SetSizer(bSizer_resultados)
        self.m_panel_resultados.Layout()
        bSizer_resultados.Fit(self.m_panel_resultados)
        bSizer_panelResultados.Add(self.m_panel_resultados, 1, wx.EXPAND | wx.ALL, 5)

        bSizer_principal.Add(bSizer_panelResultados, 1, wx.EXPAND, 5)

        bSizer171 = wx.BoxSizer(wx.HORIZONTAL)

        self.lbl_nota1 = wx.StaticText(self, wx.ID_ANY, u"// DobleClick sobre el encabezado de la tabla para ordenar",
                                       wx.DefaultPosition, wx.DefaultSize, 0)
        self.lbl_nota1.Wrap(-1)
        self.lbl_nota1.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 90, False, wx.EmptyString))
        self.lbl_nota1.SetForegroundColour(wx.Colour(0, 128, 0))

        bSizer171.Add(self.lbl_nota1, 1, wx.ALL, 5)

        self.btn_exportar_xlsx = wx.Button(self, wx.ID_ANY, u"-->&XLSX", wx.DefaultPosition, wx.DefaultSize,
                                           wx.NO_BORDER)
        self.btn_exportar_xlsx.SetForegroundColour(wx.Colour(0, 0, 0))
        self.btn_exportar_xlsx.SetBackgroundColour(wx.Colour(128, 255, 0))

        bSizer171.Add(self.btn_exportar_xlsx, 0, wx.ALL, 5)

        self.btn_exportat_pdf = wx.Button(self, wx.ID_ANY, u"-->&PDF", wx.DefaultPosition, wx.DefaultSize, wx.NO_BORDER)
        self.btn_exportat_pdf.SetForegroundColour(wx.Colour(255, 255, 255))
        self.btn_exportat_pdf.SetBackgroundColour(wx.Colour(223, 0, 112))

        bSizer171.Add(self.btn_exportat_pdf, 0, wx.ALL, 5)

        bSizer_principal.Add(bSizer171, 0, wx.EXPAND, 5)

        self.SetSizer(bSizer_principal)

        ## VALORES INICIALES EAY
        self.func_valoresIniciales()


        self.Layout()

        self.Centre(wx.BOTH)

        # Connect Events
        self.radioBox_area.Bind(wx.EVT_RADIOBOX, self.radioBox_areaOnRadioBox)
        self.radioBox_periodo.Bind(wx.EVT_RADIOBOX, self.radioBox_periodoOnRadioBox)
        self.btn_buscar.Bind(wx.EVT_BUTTON, self.btn_buscarOnButtonClick)
        self.btn_graficar.Bind(wx.EVT_BUTTON, self.btn_graficarOnButtonClick)
        self.grid_resultado_busqueda.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK,
                                          self.grid_resultado_busquedaOnGridCellLeftDClick)
        self.grid_estadisticas.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, self.grid_resultado_busquedaOnGridCellLeftDClick)
        self.btn_exportar_xlsx.Bind(wx.EVT_BUTTON, self.btn_exportar_xlsxOnButtonClick)
        self.btn_exportat_pdf.Bind(wx.EVT_BUTTON, self.btn_exportat_pdfOnButtonClick)

        self.grid_resultado_busqueda.Bind(wx.grid.EVT_GRID_LABEL_LEFT_DCLICK,
                                          self.grid_resultado_busquedaOnGridLabelLeftDClick)

    def __del__(self):
        pass

    def grid_resultado_busquedaOnGridLabelLeftDClick(self, event):

        columna = event.GetCol()

        if self.orden_ascendente == True:
            self.orden_ascendente = False
        else:
            self.orden_ascendente = True

        self.func_ordenarGrillaPorColumna(columna, self.orden_ascendente)

        self.colorearGrilla()

        event.Skip()

    # Virtual event handlers, overide them in your derived class
    def radioBox_areaOnRadioBox(self, event):
        area_produccion = self.radioBox_area.GetStringSelection()
        self.cargar_checkList_turno(area_produccion)
        event.Skip()

    def radioBox_periodoOnRadioBox(self, event):
        event.Skip()

    def btn_buscarOnButtonClick(self, event):
        from formEAY.utilCAC.Utiles_proposito_general import ManejoFechasHoras

        turnos = []
        for i in self.checkList_turno.GetCheckedStrings():
            turnos.append(i)

        self.lbl_turnos.SetLabel(str(turnos))

        area_produccion = self.radioBox_area.GetStringSelection()
        periodo = self.radioBox_periodo.GetStringSelection()

        cad_turno = ''
        lista_turnos_sel = self.checkList_turno.GetCheckedStrings()

        if area_produccion != 'DESPACHOS':
            if len(lista_turnos_sel) == 0:
                wx.MessageBox(u'Debes seleccionar mínimo un Turno', u'Atención', wx.OK | wx.ICON_INFORMATION)
                return 0

            for i in lista_turnos_sel:
                el_id = self.dic_turnos_todos[i][0]
                cad_turno += str(el_id) + ','
            cad_turno = cad_turno[:-1]

            if cad_turno != '':
                cad_turno = ' and id_turno in (' + cad_turno + ' )'

        fecha_inicio = ManejoFechasHoras.formatearFechaXSql(self.datePicker_fecha1)
        fecha_fin = ManejoFechasHoras.formatearFechaXSql(self.datePicker_fecha2)

        if area_produccion != 'DESPACHOS':
            rows, cabeceras = self.buscar(area_produccion, fecha_inicio, fecha_fin, cad_turno, periodo)

        if area_produccion == 'DESPACHOS':
            sSql= ''

            ##-----------------------------------
            cabeceras_estadisticas = ['Total Unid', 'Unidades 1ra', 'Unid 2da']
            cabeceras_estadisticas2 = ['Total Ton', 'Ton 1ra', 'Ton 2da']

            if periodo == 'DIA':
                sSql = """  SELECT  dm.fecha, sum(dp.cant_primera + dp.cant_segunda) as total_unids, sum(dp.cant_primera) as primera, sum(dp.cant_segunda) as segunda, 
                                        trunc((sum(dp.cant_primera)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_primera,
                                        trunc((sum(dp.cant_segunda)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_segunda,
                                        trunc(sum(p.peso * (dp.cant_primera + dp.cant_segunda))/1000000.0, 2) as total_ton,
                                        trunc(sum(p.peso * (dp.cant_primera))/1000000.0, 2) as ton_primera,
                                        trunc(sum(p.peso * (dp.cant_segunda))/1000000.0, 2) as ton_segunda
                            FROM detalle_despacho as dp, producto as p, despacho_mercancia as dm  
                            WHERE dp.activo = true and p.id_producto = dp.id_producto and dm.uuid = dp.uuid and
                                    dm.fecha >= '{0}' and dm.fecha <= '{1}'
                            GROUP BY dm.fecha
                            ORDER BY dm.fecha DESC        
                                                                    """.format(fecha_inicio, fecha_fin)

                cabeceras = ['Fecha', 'Total Unid', 'Unid 1ra', 'Unid 2da', '% 1ra', '% 2da',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
                cols_estadisticas = [1, 2, 3]
                cols_estadisticas2 = [6, 7, 8]

            if periodo == 'AÑO':
                sSql = """  SELECT  date_part('year', dm.fecha) AS anyo, sum(dp.cant_primera + dp.cant_segunda) as total_unids, sum(dp.cant_primera) as primera, sum(dp.cant_segunda) as segunda, 
                                    trunc((sum(dp.cant_primera)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_primera,
                                    trunc((sum(dp.cant_segunda)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_segunda,
                                    trunc(sum(p.peso * (dp.cant_primera + dp.cant_segunda))/1000000.0, 2) as total_ton,
                                    trunc(sum(p.peso * (dp.cant_primera))/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso * (dp.cant_segunda))/1000000.0, 2) as ton_segunda
                                FROM  detalle_despacho as dp, producto as p, despacho_mercancia as dm  
                            WHERE dp.activo = true and p.id_producto = dp.id_producto and dm.uuid = dp.uuid and
                                    dm.fecha >= '{0}' and dm.fecha <= '{1}'
                            GROUP BY anyo
                            ORDER BY anyo DESC                             
                                                """.format(fecha_inicio, fecha_fin)

                cabeceras = [u'Año', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
                cols_estadisticas = [1, 2, 3]
                cols_estadisticas2 = [6, 7, 8]

            if periodo == 'SEMANA':
                sSql = """  SELECT date_part('year', dm.fecha) as anyo, date_part('week', dm.fecha) as semana,
                                    sum(dp.cant_primera + dp.cant_segunda) as total_unids, sum(dp.cant_primera) as primera, sum(dp.cant_segunda) as segunda, 
                                    trunc((sum(dp.cant_primera)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_primera,
                                    trunc((sum(dp.cant_segunda)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_segunda,
                                    trunc(sum(p.peso * (dp.cant_primera + dp.cant_segunda))/1000000.0, 2) as total_ton,
                                    trunc(sum(p.peso * (dp.cant_primera))/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso * (dp.cant_segunda))/1000000.0, 2) as ton_segunda
                            FROM  detalle_despacho as dp, producto as p, despacho_mercancia as dm  
                            WHERE dp.activo = true and p.id_producto = dp.id_producto and dm.uuid = dp.uuid and
                                    dm.fecha >= '{0}' and dm.fecha <= '{1}'
                            GROUP BY anyo, semana
                            ORDER BY anyo, semana DESC    
                                                                    """.format(fecha_inicio, fecha_fin)

                cabeceras = [u'Año', 'Semana', 'Total Unid', 'Unid 1ra', 'Unid 2da', '% 1ra', '% 2da',
                             'Total Ton', 'Ton 1ra', 'Ton 2da']
                cols_estadisticas = [2, 3, 4]
                cols_estadisticas2 = [7, 8, 9]

            if periodo == 'MES':
                sSql = """  SELECT  date_part('year', dm.fecha) as anyo, date_part('month', dm.fecha) as mes,
                                    sum(dp.cant_primera + dp.cant_segunda) as total_unids, sum(dp.cant_primera) as primera, sum(dp.cant_segunda) as segunda, 
                                    trunc((sum(dp.cant_primera)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_primera,
                                    trunc((sum(dp.cant_segunda)/(sum(dp.cant_primera + dp.cant_segunda))* 100.0), 2) as porc_segunda,
                                    trunc(sum(p.peso * (dp.cant_primera + dp.cant_segunda))/1000000.0, 2) as total_ton,
                                    trunc(sum(p.peso * (dp.cant_primera))/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso * (dp.cant_segunda))/1000000.0, 2) as ton_segunda
                            FROM  detalle_despacho as dp, producto as p, despacho_mercancia as dm  
                            WHERE dp.activo = true and p.id_producto = dp.id_producto and dm.uuid = dp.uuid and
                                    dm.fecha >= '{0}' and dm.fecha <= '{1}'
                            GROUP BY anyo, mes
                            ORDER BY anyo, mes DESC       
                                                                    """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Mes', 'Total Unid', 'Unid 1ra', 'Unid 2da', '% 1ra', '% 2da',
                             'Total Ton', 'Ton 1ra', 'Ton 2da']
                cols_estadisticas = [2, 3, 4]
                cols_estadisticas2 = [7, 8, 9]
            ##-----------------------------------

            cabeceras = []
            rows = Ejecutar_SQL.select_varios_registros(sSql, 'frm_report_unidades_por_periodo/btn_buscarOnButtonClick',
                                                                            1000, BasesDeDatos.DB_PRINCIPAL)

            if rows == None:
                self.m_panel_resultados.Hide()
            else:
                self.lbl_area.SetLabel('DESPACHOS')

                cad_fecha = fecha_inicio + '  al  ' + fecha_fin
                self.lbl_fechas.SetLabel(cad_fecha)
                cant_registros = str(len(rows))
                #self.lbl_registrosencontrados.SetLabel(cant_registros)

                list_cabeceras = []

                if periodo == 'DIA':
                    ManipularGrillas.setCantidadColumnasGrilla(self.grid_resultado_busqueda, 9)
                    list_cabeceras = ['Fecha', 'Total unids', 'Primera', 'Segunda', '% Primera', '% Segunda',
                                      'Total ton', 'Ton Primera', 'Ton Segunda']
                if periodo == 'AÑO':
                    ManipularGrillas.setCantidadColumnasGrilla(self.grid_resultado_busqueda, 9)
                    list_cabeceras = ['Fecha', 'Total unids', 'Primera', 'Segunda', '% Primera', '% Segunda',
                                      'Total ton', 'Ton Primera', 'Ton Segunda']
                if periodo == 'SEMANA':
                    ManipularGrillas.setCantidadColumnasGrilla(self.grid_resultado_busqueda, 10)
                    list_cabeceras = [u'Año', u'Semana', 'Total unids', 'Primera', 'Segunda', '% Primera', '% Segunda',
                                      'Total ton', 'Ton Primera', 'Ton Segunda']
                if periodo == 'MES':
                    ManipularGrillas.setCantidadColumnasGrilla(self.grid_resultado_busqueda, 10)
                    list_cabeceras = [u'Año', u'Mes', 'Total unids', 'Primera', 'Segunda', '% Primera', '% Segunda',
                                      'Total ton', 'Ton Primera', 'Ton Segunda']
                ManipularGrillas.setCabecerasGrilla(self.grid_resultado_busqueda, list_cabeceras)


                ##------------------------------------------
                ## para estadisticas de despachos

                # rows_matriz = pd.DataFrame(rows, columns = list_cabeceras)
                #
                # sumatoria = rows_matriz[['Total unids', 'Primera', 'Segunda', 'Total ton', 'Ton Primera', 'Ton Segunda']].sum()
                # media = rows_matriz[['Total unids', 'Primera', 'Segunda', 'Total ton', 'Ton Primera', 'Ton Segunda']].mean()
                # desviacion = rows_matriz.std(axis=0)
                # mediana = rows_matriz[['Total unids', 'Primera', 'Segunda', 'Total ton', 'Ton Primera', 'Ton Segunda']].median()
                # maximo = rows_matriz[['Total unids', 'Primera', 'Segunda', 'Total ton', 'Ton Primera', 'Ton Segunda']].max()
                # minimo = rows_matriz[['Total unids', 'Primera', 'Segunda', 'Total ton', 'Ton Primera', 'Ton Segunda']].min()
                #
                # print('la desviacion', desviacion)
                #
                # for j in range(3):
                #     self.grid_estadisticas.SetCellValue(0, j, str(sumatoria[j]))
                #     self.grid_estadisticas2.SetCellValue(0, j, str(sumatoria[j+3]))
                #
                #     self.grid_estadisticas.SetCellValue(1, j, str(media[j]))
                #     self.grid_estadisticas2.SetCellValue(1, j, str(media[j + 3]))
                #
                #     self.grid_estadisticas.SetCellValue(2, j, str(desviacion[j]))
                #     self.grid_estadisticas2.SetCellValue(2, j, str(desviacion[j + 3]))
                ##-------------------------------------------

                self.m_panel_resultados.Show()


        if rows == None:
            self.m_panel_resultados.Hide()
            self.btn_graficar.Hide()
            self.btn_exportar_xlsx.Hide()

        else:
            self.lbl_area.SetLabel(area_produccion)

            cad_fecha = fecha_inicio + '  al  ' + fecha_fin
            self.lbl_fechas.SetLabel(cad_fecha)
            self.lbl_periodo.SetLabel(periodo)

            self.m_panel_resultados.Show()
            self.btn_graficar.Show()
            self.btn_exportar_xlsx.Show()

        ManipularGrillas.llenarGrilla(self.grid_resultado_busqueda, rows)
        ManipularGrillas.reemplazarValorCeldaGrilla(self.grid_resultado_busqueda, 'None', '0')

        if periodo in ['SEMANA', 'MES']:
            self.redondear_columna_grilla([0, 1], 0)
        if periodo in [u'AÑO']:
            self.redondear_columna_grilla([0], 0)
        if area_produccion == AREA_CARGUE_VAGONETAS:
            if periodo in [u'AÑO', 'DIA']:
                self.redondear_columna_grilla([4, 5], 2)
            else:
                self.redondear_columna_grilla([5, 6], 2)

        self.colorearGrilla()
        self.Layout()
        event.Skip()

    def btn_graficarOnButtonClick(self, event):

        cant_filas = self.grid_resultado_busqueda.GetNumberRows()

        if cant_filas == 0:
            wx.MessageBox(u'No hay datos para graficar, intentalo presionando el Boton Buscar', u'Atención',
                          wx.OK | wx.ICON_INFORMATION)
            return 0

        area = self.lbl_area.GetLabel()
        fechas = self.lbl_fechas.GetLabel()
        turnos = self.lbl_turnos.GetLabel()
        periodo = self.lbl_periodo.GetLabel()
        columna = 0

        orden_ascendente = 0
        if periodo in ('MES', u'AÑO'):
            columna = 0
        if periodo in ('SEMANA', u'MES'):
            columna = 1
        #self.func_ordenarGrillaPorColumna(columna, orden_ascendente)

        #self.colorearGrilla()

        dic_valores_resumen = {}
        dic_valores_detalle = {}

        dic_cabecera = {'area': area, 'fechas': fechas, 'turnos': turnos, 'periodo': periodo}
        dic_valores = {}

        list_unidades = []
        list_fechas = []

        list_primera_unidades = []
        list_rotura_unidades = []
        list_primera_ton = []
        list_rotura_ton = []
        list_segunda_unidades = []
        list_segunda_ton = []

        maximo_unidades = int(self.grid_estadisticas.GetCellValue(4, 0))
        maximo_toneladas = int(float(self.grid_estadisticas2.GetCellValue(4, 0)))
        n_datos = int(self.grid_estadisticas2.GetCellValue(6, 0))

        if area == 'EXTRUSION':
            if periodo in ('DIA', u'AÑO'):
                list_fechas = self.rows_matriz[:, 0]
                list_primera_unidades = self.rows_matriz[:, 1]
                list_primera_ton = self.rows_matriz[:, 2]
            if periodo in ('SEMANA', 'MES'):
                list_fechas = self.rows_matriz[:, 1]
                list_primera_unidades = self.rows_matriz[:, 2]
                list_primera_ton = self.rows_matriz[:, 3]

            dic_valores_detalle = {'x': list_fechas, 'y_u': list_primera_unidades,
                                   'y_t': list_primera_ton, }

        if area == 'CARGUE DE VAGONETAS':
            if periodo in ('DIA', u'AÑO'):
                list_fechas = self.rows_matriz[:, 0]
                list_primera_unidades = self.rows_matriz[:, 2]
                list_rotura_unidades = self.rows_matriz[:, 3]
                list_primera_ton = self.rows_matriz[:, 7]
                list_rotura_ton = self.rows_matriz[:, 8]
            if periodo in ('SEMANA', 'MES'):
                list_fechas = self.rows_matriz[:, 1]
                list_primera_unidades = self.rows_matriz[:, 3]
                list_rotura_unidades = self.rows_matriz[:, 4]
                list_primera_ton = self.rows_matriz[:, 8]
                list_rotura_ton = self.rows_matriz[:, 9]

            dic_valores_detalle = {'x': list_fechas,    'y1_u': list_primera_unidades,  'y2_u': list_rotura_unidades,
                                                        'y1_t': list_primera_ton,  'y2_t': list_rotura_ton }


        if area == 'DESCARGUE DE VAGONETAS':
            if periodo in ('DIA', u'AÑO'):
                list_fechas = self.rows_matriz[:, 0]
                list_primera_unidades = self.rows_matriz[:, 2]
                list_segunda_unidades = self.rows_matriz[:, 3]
                list_rotura_unidades = self.rows_matriz[:, 4]
                list_primera_ton = self.rows_matriz[:, 9]
                list_segunda_ton = self.rows_matriz[:, 10]
                list_rotura_ton = self.rows_matriz[:, 11]
            if periodo in ('SEMANA', 'MES'):
                list_fechas = self.rows_matriz[:, 1]
                list_primera_unidades = self.rows_matriz[:, 3]
                list_segunda_unidades = self.rows_matriz[:, 4]
                list_rotura_unidades = self.rows_matriz[:, 5]
                list_primera_ton = self.rows_matriz[:, 10]
                list_segunda_ton = self.rows_matriz[:, 11]
                list_rotura_ton = self.rows_matriz[:, 12]

            dic_valores_detalle = {'x': list_fechas,    'y1_u': list_primera_unidades, 'y2_u': list_segunda_unidades, 'y3_u': list_rotura_unidades,
                                                        'y1_t': list_primera_ton, 'y2_t': list_segunda_ton, 'y3_t': list_rotura_ton}

        import formEAY.formularios.frm_graficas.frm_graficosProduccionPorPeriodo as frm_graficosProduccionPorPeriodo
        frame_graficosProduccionPorPeriodo = frm_graficosProduccionPorPeriodo.GraficosProduccionPorPeriodo(self, dic_cabecera, dic_valores_detalle,
                                                                                                           maximo_unidades, maximo_toneladas, n_datos)
        frame_graficosProduccionPorPeriodo.Center()
        frame_graficosProduccionPorPeriodo.Show()

        event.Skip()

    def grid_resultado_busquedaOnGridCellLeftDClick(self, event):
        event.Skip()

    def btn_exportar_xlsxOnButtonClick(self, event):
        from openpyxl import Workbook
        from openpyxl.styles import colors
        from openpyxl.styles import Font, Color
        from openpyxl.styles import PatternFill
        import os
        from datetime import datetime
        import shutil

        wb = Workbook()
        # grab the active worksheet
        ws = wb.active

        relleno = PatternFill(start_color='FFAFEEEE', end_color='FFAFEEEE', fill_type='solid')  # 'FFAFEEEE'

        ft_titulo = Font(color="BF0041", size=14, bold=True)
        ft_cabecera = Font(color="BF0041", size=12, bold=True)

        ws.title = "Produccion Por Periodo"

        ##   -------------------------------
        area_produccion =self.lbl_area.GetLabel()
        periodo = self.lbl_periodo.GetLabel()
        if area_produccion == AREA_CARGUE_VAGONETAS:
            if periodo == 'DIA':
                cabeceras = ['Fecha', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']

            if periodo == 'AÑO':
                cabeceras = [u'Año', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
            if periodo == 'SEMANA':
                cabeceras = [u'Año', 'Semana', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
            if periodo == 'MES':
                cabeceras = [u'Año', 'Mes', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
        if area_produccion == AREA_DESCARGUE_VAGONETAS:
            if periodo == 'DIA':
                cabeceras = ['Fecha', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
            if periodo == 'SEMANA':
                cabeceras = [u'Año', 'Semana', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da',
                             '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
            if periodo == 'MES':
                cabeceras = [u'Año', 'Mes', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da',
                             '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
            if periodo == 'AÑO':
                cabeceras = [u'Año', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
        if area_produccion == AREA_EXTRUSION:
            if periodo == 'DIA':
                cabeceras = ['Fecha', 'Unidades', 'Toneladas']
            if periodo == 'SEMANA':
                cabeceras = [u'año', 'Semana', 'Unidades', 'Toneladas']
            if periodo == 'MES':
                cabeceras = [u'año', 'Mes', 'Unidades', 'Toneladas']
            if periodo == u'AÑO':
                cabeceras = [u'Año', 'Unidades', 'Toneladas']
        ##  --------------------------------

        nomCarpeta = 'REPORTES\XLSX\PRODUCCCION POR PERIODO'
        nomArchivo = 'Produccion por periodo'

        try:
            dir_trabajo = os.getcwd()
            rutaCarpeta = os.path.join(dir_trabajo, nomCarpeta)

            if not os.path.exists(nomCarpeta):
                os.mkdir(nomCarpeta)
            la_ruta = os.path.join(dir_trabajo, nomCarpeta)

            ahora = datetime.now()

            la_fecha = str(ahora.year) + '-' + str(ahora.month) + '-' + str(ahora.day) + '  HH--' + str(
                ahora.hour) + '-' + str(ahora.minute) + '-' + str(ahora.second)

            nombreCompletoArchivo = nomArchivo + ' ' + la_fecha + u'.xlsx'

            # Data can be assigned directly to cells
            # sheet['A1'] = 42
            ws['A1'] = nomArchivo
            ws['A2'] = 'Fecha:'
            ws['A3'] = 'Hora:'
            ws['B2'] = ahora.date()
            ws['B3'] = ahora.time()

            ws['A5'] = 'Area:'
            ws['A6'] = 'Rango fechas:'
            ws['A7'] = 'Periodo:'
            ws['A8'] = 'Turnos:'

            ws['A9'] = ''  # filas vacias pora dar orden
            ws['A10'] = ''

            ws['B5'] = self.lbl_area.GetLabel()
            ws['B6'] = self.lbl_fechas.GetLabel()
            ws['B7'] = self.lbl_periodo.GetLabel()
            ws['B8'] = self.lbl_turnos.GetLabel()

            titulo = ws['A1']
            titulo.font = ft_titulo

            ws.merge_cells('A1:I1')





            ws.merge_cells('B2:I2')
            ws.merge_cells('B3:I3')

            ws.merge_cells('A4:I4')

            ws.merge_cells('B5:I5')
            ws.merge_cells('B6:I6')
            ws.merge_cells('B7:I7')
            ws.merge_cells('B8:I8')

            ws.merge_cells('A9:I9')
            ws.merge_cells('A10:I10')

            ws.append(cabeceras)
            rows_datos = self.rows_matriz.tolist()

            for fila in rows_datos:
                ws.append(fila)

            for j in range(len(cabeceras)):
                cell_range = ws.cell(row=11, column=j+1)  # era row=1
                cell_range.fill = relleno

            cell_range = ws.cell(row=1, column=1)  # era row=1
            cell_range.fill = relleno

            from openpyxl.styles import Alignment
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['B2'].alignment = Alignment(horizontal="left")
            ws['B3'].alignment = Alignment(horizontal="left")

            wb.save(nombreCompletoArchivo)

            shutil.move(nombreCompletoArchivo, la_ruta)

            wx.MessageBox(u'Se ha realizado con exito la Exportación', u'Atención', wx.OK | wx.ICON_INFORMATION)

            return nombreCompletoArchivo
        except:
            return 0


        event.Skip()

    def btn_exportat_pdfOnButtonClick(self, event):
        event.Skip()

    ##  FUNCIONES EAY



    def func_ordenarGrillaPorColumna(self, columna, orden_ascendente):
        area = self.radioBox_area.GetStringSelection()
        periodo = self.radioBox_periodo.GetStringSelection()
        lista_tipoColumna = []

        if area == 'DESPACHOS':
            if periodo in ('DIA', u'AÑO'):
                lista_tipoColumna = ['str', 'int', 'int', 'int', 'float', 'float', 'float', 'float', 'float']
            if periodo in ('SEMANA', u'MES'):
                lista_tipoColumna = ['str', 'int', 'int', 'int', 'int', 'float', 'float', 'float', 'float', 'float']

        if area == 'EXTRUSION':
            if periodo in ('DIA', u'AÑO'):
                lista_tipoColumna = ['str', 'int', 'float']
            if periodo in ('SEMANA', u'MES'):
                lista_tipoColumna = ['str', 'int', 'int', 'float']
        if area == 'CARGUE DE VAGONETAS':
            if periodo in ('DIA', u'AÑO'):
                lista_tipoColumna = ['str', 'int', 'int', 'int', 'float', 'float', 'float', 'float', 'float']
            if periodo in ('SEMANA', u'MES'):
                lista_tipoColumna = ['str', 'int', 'int', 'int', 'int', 'float', 'float', 'float', 'float', 'float']
        if area == 'DESCARGUE DE VAGONETAS':
            if periodo in ('DIA', u'AÑO'):
                lista_tipoColumna = ['str', 'int', 'int', 'int', 'int', 'float', 'float', 'float', 'float', 'float',
                                     'float', 'float']
            if periodo in ('SEMANA', u'MES'):
                lista_tipoColumna = ['str', 'int', 'int', 'int', 'int', 'int', 'float', 'float', 'float', 'float',
                                     'float', 'float', 'float']

        ManipularGrillas.ordenarGrillaPorColumna(self.grid_resultado_busqueda, columna, lista_tipoColumna,
                                                 orden_ascendente)

    def func_valoresIniciales(self):
        self.btn_exportat_pdf.Hide()  # FUNCION NO PROGRAMADA

        self.m_panel_resultados.Hide()
        self.btn_graficar.Hide()
        self.btn_exportar_xlsx.Hide()

        self.grid_resultado_busqueda.AutoSizeColumns()
        self.grid_estadisticas.AutoSizeColumns()
        self.grid_estadisticas2.AutoSizeColumns()
        self.cargar_checkList_turno('EXTRUSION')


    def colorearGrilla(self):
        area_produccion = self.lbl_area.GetLabel()
        periodo = self.lbl_periodo.GetLabel()

        if area_produccion == AREA_EXTRUSION:
            if periodo in ('DIA', u'AÑO'):
                dic_color = {1: COLOR_RESALTE_PRIMERA, 2: COLOR_RESALTE_SEGUNDA}
            if periodo in ('SEMANA', u'MES'):
                dic_color = {2: COLOR_RESALTE_PRIMERA, 3: COLOR_RESALTE_SEGUNDA}
            ManipularGrillas.setColorFondoCeldaGrilla(self.grid_resultado_busqueda, dic_color)

        if area_produccion == AREA_CARGUE_VAGONETAS:
            if periodo in ('DIA', u'AÑO'):
                dic_color = {2: COLOR_RESALTE_TOTAL, 3: COLOR_RESALTE_ROTURA_CLARO
                             }
            if periodo in ('SEMANA', u'MES'):
                dic_color = {3: COLOR_RESALTE_TOTAL, 4: COLOR_RESALTE_ROTURA_CLARO
                             }

            ManipularGrillas.setColorFondoCeldaGrilla(self.grid_resultado_busqueda, dic_color)

        if area_produccion == AREA_DESCARGUE_VAGONETAS:
            if periodo in ('DIA', u'AÑO'):
                dic_color = {2: COLOR_RESALTE_TOTAL, 3: COLOR_RESALTE_SEGUNDA, 4: COLOR_RESALTE_ROTURA
                             }
            if periodo in ('SEMANA', u'MES'):
                dic_color = {3: COLOR_RESALTE_TOTAL, 4: COLOR_RESALTE_SEGUNDA, 5: COLOR_RESALTE_ROTURA
                             }

        if area_produccion == 'DESPACHOS':
            if periodo in ('DIA', u'AÑO'):
                dic_color = {2: COLOR_RESALTE_TOTAL, 3: COLOR_RESALTE_SEGUNDA
                             }
            if periodo in ('SEMANA', u'MES'):
                dic_color = {3: COLOR_RESALTE_TOTAL, 4: COLOR_RESALTE_SEGUNDA
                             }


        ManipularGrillas.setColorFondoCeldaGrilla(self.grid_resultado_busqueda, dic_color)


        self.Layout()


    def buscar(self, area_produccion, fecha_inicio, fecha_fin, cad_turno, periodo):
        rows = []
        sSql = ''
        cabeceras = []
        cols_estadisticas = []
        cabeceras_estadisticas = []

        if area_produccion == AREA_CARGUE_VAGONETAS:
            cabeceras_estadisticas = ['Total Unid', 'Unidades 1ra', 'Unid Rotura']
            cabeceras_estadisticas2 = ['Total Ton', 'Ton 1ra', 'Ton Rotura']

            if periodo == 'DIA':
                sSql = """  SELECT tablaA.fecha, tablaA.primera AS PRIMERA, tablaB.rotura AS ROTURA,
                                    trunc(tablaA.ton_primera, 2) as ton_primera,  trunc(tablaB.ton_rotura, 2) as ton_rotura
                            FROM 	(
                                    SELECT cp_ecd.fecha_inicio as fecha, sum(dcv.unidades_producto) AS primera,
                                            sum(dcv.unidades_producto * p.peso /1000000.0) as ton_primera
                                    FROM  cabecera_proceso_ecd cp_ecd, detalle_cargue_vagonetas as dcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}'
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and dcv.uuid=cp_ecd.uuid
                                            and p.id_producto = dcv.id_producto
                                            and cp_ecd.activo = 'True' {2}
                                    GROUP BY cp_ecd.fecha_inicio
                                    ) as tablaA
                            FULL OUTER JOIN 
                                    (
                                    SELECT cp_ecd.fecha_inicio as fecha, COALESCE(sum(rcv.cant_rotos), '0') AS rotura,
                                            sum(rcv.cant_rotos * p.peso /1000000.0) as ton_rotura
                                    FROM  cabecera_proceso_ecd cp_ecd, rotura_cargue_vagonetas as rcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}' 
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and rcv.uuid=cp_ecd.uuid
                                            and p.id_producto = rcv.id_producto
                                            and cp_ecd.activo = 'True'  {2} 
                                    GROUP BY cp_ecd.fecha_inicio
                                    ) as tablaB
                            ON   tablaA.fecha = tablaB.fecha 
                            ORDER BY  tablaA.fecha         
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = ['Fecha', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
                cols_estadisticas = [1, 2, 3]
                cols_estadisticas2 = [6, 7, 8]

            if periodo == 'AÑO':
                sSql = """  SELECT tablaA.fecha, tablaA.primera AS PRIMERA, tablaB.rotura AS ROTURA,
                                    trunc(tablaA.ton_primera, 2) as ton_primera,  trunc(tablaB.ton_rotura, 2) as ton_rotura
                            FROM 	(
                                    SELECT date_part('year', cp_ecd.fecha_inicio) as fecha, sum(dcv.unidades_producto) AS primera,
                                            sum(dcv.unidades_producto * p.peso /1000000.0) as ton_primera
                                    FROM  cabecera_proceso_ecd cp_ecd, detalle_cargue_vagonetas as dcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}'
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and dcv.uuid=cp_ecd.uuid
                                            and p.id_producto = dcv.id_producto
                                            and cp_ecd.activo = 'True' {2}
                                    GROUP BY fecha
                                    ) as tablaA
                            FULL OUTER JOIN 
                                    (
                                    SELECT date_part('year', cp_ecd.fecha_inicio) as fecha, SUM(rcv.cant_rotos) AS rotura,				
                                            sum(rcv.cant_rotos * p.peso /1000000.0) as ton_rotura
                                    FROM  cabecera_proceso_ecd cp_ecd, rotura_cargue_vagonetas as rcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}' 
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and rcv.uuid=cp_ecd.uuid
                                            and p.id_producto = rcv.id_producto
                                            and cp_ecd.activo = 'True'  {2} 
                                    GROUP BY fecha
                                    ) as tablaB
                            ON   tablaA.fecha = tablaB.fecha    
                            ORDER BY  tablaA.fecha          
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
                cols_estadisticas = [1, 2, 3]
                cols_estadisticas2 = [6, 7, 8]

            if periodo == 'SEMANA':
                sSql = """  SELECT 	tablaA.anyo, tablaA.semana, 
                                    tablaA.primera AS PRIMERA, tablaB.rotura AS ROTURA,
                                    trunc(tablaA.ton_primera, 2) as ton_primera,  trunc(tablaB.ton_rotura, 2) as ton_rotura
                            FROM 	(
                                    SELECT date_part('year', cp_ecd.fecha_inicio) as anyo, date_part('week', cp_ecd.fecha_inicio) as semana, 
                                            sum(dcv.unidades_producto) AS primera,
                                            sum(dcv.unidades_producto * p.peso /1000000.0) as ton_primera
                                    FROM  cabecera_proceso_ecd cp_ecd, detalle_cargue_vagonetas as dcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}' 
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and dcv.uuid=cp_ecd.uuid
                                            and p.id_producto = dcv.id_producto
                                            and cp_ecd.activo = 'True' {2} 
                                    GROUP BY anyo, semana
                                    ) as tablaA
                            FULL OUTER JOIN 
                                    (
                                    SELECT date_part('year', cp_ecd.fecha_inicio) as anyo, date_part('week', cp_ecd.fecha_inicio) as semana, 
                                            SUM(rcv.cant_rotos) AS rotura,				
                                            sum(rcv.cant_rotos * p.peso /1000000.0) as ton_rotura
                                    FROM  cabecera_proceso_ecd cp_ecd, rotura_cargue_vagonetas as rcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}' 
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and rcv.uuid=cp_ecd.uuid
                                            and p.id_producto = rcv.id_producto
                                            and cp_ecd.activo = 'True' {2} 
                                    GROUP BY anyo, semana
                                    ) as tablaB
                            ON   tablaA.anyo = tablaB.anyo and tablaA.semana = tablaB.semana   
                            ORDER BY tablaA.anyo, tablaA.semana       
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Semana', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
                cols_estadisticas = [2, 3, 4]
                cols_estadisticas2 = [7, 8, 9]

            if periodo == 'MES':
                sSql = """  SELECT 	tablaA.anyo, tablaA.mes, 
                                    tablaA.primera AS PRIMERA, tablaB.rotura AS ROTURA,
                                    trunc(tablaA.ton_primera, 2) as ton_primera,  trunc(tablaB.ton_rotura, 2) as ton_rotura
                            FROM 	(
                                    SELECT date_part('year', cp_ecd.fecha_inicio) as anyo, date_part('month', cp_ecd.fecha_inicio) as mes, 
                                            sum(dcv.unidades_producto) AS primera,
                                            sum(dcv.unidades_producto * p.peso /1000000.0) as ton_primera
                                    FROM  cabecera_proceso_ecd cp_ecd, detalle_cargue_vagonetas as dcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}' 
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and dcv.uuid=cp_ecd.uuid
                                            and p.id_producto = dcv.id_producto
                                            and cp_ecd.activo = 'True' {2} 
                                    GROUP BY anyo, mes
                                    ) as tablaA
                            FULL OUTER JOIN 
                                    (
                                    SELECT date_part('year', cp_ecd.fecha_inicio) as anyo, date_part('month', cp_ecd.fecha_inicio) as mes, 
                                            SUM(rcv.cant_rotos) AS rotura,				
                                            sum(rcv.cant_rotos * p.peso /1000000.0) as ton_rotura
                                    FROM  cabecera_proceso_ecd cp_ecd, rotura_cargue_vagonetas as rcv, producto as p
                                    WHERE   cp_ecd.fecha_inicio >= '{0}'  and cp_ecd.fecha_inicio <= '{1}' 
                                            and cp_ecd.area_produccion = 'CARGUE DE VAGONETAS'  and rcv.uuid=cp_ecd.uuid
                                            and p.id_producto = rcv.id_producto
                                            and cp_ecd.activo = 'True' {2} 
                                    GROUP BY anyo, mes
                                    ) as tablaB
                            ON   tablaA.anyo = tablaB.anyo and tablaA.mes = tablaB.mes  
                            ORDER BY tablaA.anyo, tablaA.mes        
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Mes', 'Total Unid', 'Unid 1ra', 'Unid Rotura', '% 1ra', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton Rotura']
                cols_estadisticas = [2, 3, 4]
                cols_estadisticas2 = [7, 8, 9]

        if area_produccion == AREA_DESCARGUE_VAGONETAS:
            cabeceras_estadisticas = ['Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura']
            cabeceras_estadisticas2 = ['Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
            if periodo == 'DIA':
                sSql = """  SELECT c_ecd.fecha_inicio, sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos) as Total_unids,
                                    sum(de.de_primera) as primera, sum(de.de_segunda) as segunda, sum(de.rotos) as rotos,
                                    
                                    trunc(sum(de.de_primera)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_primera,
                                    trunc(sum(de.de_segunda)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_segunda, 
                                    trunc(sum(de.rotos)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_rotos,		
                                    
                                    trunc(sum(p.peso*de.de_primera)/1000000.0 + sum(p.peso*de.de_segunda)/1000000.0 + sum(p.peso*de.rotos)/1000000.0, 2) as total_ton,
                                    trunc(sum(p.peso*de.de_primera)/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso*de.de_segunda)/1000000.0, 2) as ton_segunda,
                                    trunc(sum(p.peso*de.rotos)/1000000.0, 2) as ton_rotura
                            FROM detalle_descargue_vagonetas as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2} 
                            GROUP BY de.uuid, c_ecd.fecha_inicio    
                            ORDER BY  c_ecd.fecha_inicio              
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = ['Fecha', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da', '% Rotura', 'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
                cols_estadisticas = [1, 2, 3, 4]
                cols_estadisticas2 = [8, 9, 10, 11]

            if periodo == 'SEMANA':
                sSql = """  SELECT  date_part('year',c_ecd.fecha_inicio) as anyo,  date_part('week',c_ecd.fecha_inicio) as semana,
                                    sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos) as Total_unids,
                                    sum(de.de_primera) as primera, sum(de.de_segunda) as segunda, sum(de.rotos) as rotos,

                                    trunc(sum(de.de_primera)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_primera,
                                    trunc(sum(de.de_segunda)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_segunda, 
                                    trunc(sum(de.rotos)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_rotos,		

                                    trunc(sum(p.peso*de.de_primera)/1000000.0 + sum(p.peso*de.de_segunda)/1000000.0 + sum(p.peso*de.rotos)/1000000.0, 2) as total_ton ,		
                                    trunc(sum(p.peso*de.de_primera)/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso*de.de_segunda)/1000000.0, 2) as ton_segunda,
                                    trunc(sum(p.peso*de.rotos)/1000000.0, 2) as ton_rotura
                            FROM detalle_descargue_vagonetas as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2} 
                            group by anyo,  semana    
                            ORDER BY  anyo,  semana              
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Semana', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
                cols_estadisticas = [2, 3, 4, 5]
                cols_estadisticas2 = [9, 10, 11, 12]

            if periodo == 'MES':
                sSql = """  SELECT  date_part('year',c_ecd.fecha_inicio) as anyo,  date_part('month',c_ecd.fecha_inicio) as MES,
                                    sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos) as Total_unids,
                                    sum(de.de_primera) as primera, sum(de.de_segunda) as segunda, sum(de.rotos) as rotos,

                                    trunc(sum(de.de_primera)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_primera,
                                    trunc(sum(de.de_segunda)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_segunda, 
                                    trunc(sum(de.rotos)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_rotos,		

                                    trunc(sum(p.peso*de.de_primera)/1000000.0 + sum(p.peso*de.de_segunda)/1000000.0 + sum(p.peso*de.rotos)/1000000.0, 2) as total_ton ,		
                                    trunc(sum(p.peso*de.de_primera)/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso*de.de_segunda)/1000000.0, 2) as ton_segunda,
                                    trunc(sum(p.peso*de.rotos)/1000000.0, 2) as ton_rotura
                            FROM detalle_descargue_vagonetas as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2} 
                            group by anyo,  MES    
                            ORDER BY  anyo,  MES            
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Mes', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
                cols_estadisticas = [2, 3, 4, 5]
                cols_estadisticas2 = [9, 10, 11, 12]

            if periodo == 'AÑO':
                sSql = """  SELECT  date_part('year',c_ecd.fecha_inicio) as anyo,  
                                    sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos) as Total_unids,
                                    sum(de.de_primera) as primera, sum(de.de_segunda) as segunda, sum(de.rotos) as rotos,

                                    trunc(sum(de.de_primera)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_primera,
                                    trunc(sum(de.de_segunda)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_segunda, 
                                    trunc(sum(de.rotos)*1.0 / (sum(de.de_primera) + sum(de.de_segunda)+ sum(de.rotos))*100, 2)   as porc_rotos,		

                                    trunc(sum(p.peso*de.de_primera)/1000000.0 + sum(p.peso*de.de_segunda)/1000000.0 + sum(p.peso*de.rotos)/1000000.0, 2) as total_ton ,		
                                    trunc(sum(p.peso*de.de_primera)/1000000.0, 2) as ton_primera,
                                    trunc(sum(p.peso*de.de_segunda)/1000000.0, 2) as ton_segunda,
                                    trunc(sum(p.peso*de.rotos)/1000000.0, 2) as ton_rotura
                            FROM detalle_descargue_vagonetas as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2} 
                            group by anyo    
                            ORDER BY  anyo           
                                                        """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Total Unid', 'Unid 1ra', 'Unid 2da', 'Unid Rotura', '% 1ra', '% 2da', '% Rotura',
                             'Total Ton', 'Ton 1ra', 'Ton 2da', 'Ton Rotura']
                cols_estadisticas = [1, 2, 3, 4]
                cols_estadisticas2 = [8, 9, 10, 11]

        if area_produccion == AREA_EXTRUSION:
            cabeceras_estadisticas = ['Unidades']
            cabeceras_estadisticas2 = ['Toneladas']
            if periodo == 'DIA':
                sSql = """  SELECT c_ecd.fecha_inicio, sum(de.unidades_producto), trunc(sum(p.peso * de.unidades_producto / 1000000.0), 2)
                            FROM detalle_extrusion as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2}
                            GROUP BY de.uuid, c_ecd.fecha_inicio    
                            ORDER BY  c_ecd.fecha_inicio              
                                            """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = ['Fecha', 'Unidades', 'Toneladas']
                cols_estadisticas = [1]
                cols_estadisticas2 = [2]

            if periodo == 'SEMANA':
                sSql = """  SELECT date_part('year',c_ecd.fecha_inicio) as anyo,  date_part('week',c_ecd.fecha_inicio) as semana, 
                                    sum(de.unidades_producto), trunc(sum(p.peso * de.unidades_producto / 1000000.0), 2)
                            FROM detalle_extrusion as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2}
                            group by anyo,  semana    
                            ORDER BY  anyo,  semana           
                                            """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'año', 'Semana', 'Unidades', 'Toneladas']
                cols_estadisticas = [2]
                cols_estadisticas2 = [3]
            if periodo == 'MES':
                sSql = """  SELECT date_part('year',c_ecd.fecha_inicio) as anyo,  date_part('month',c_ecd.fecha_inicio) as mes, 
                                    sum(de.unidades_producto), trunc(sum(p.peso * de.unidades_producto / 1000000.0), 2)
                            FROM detalle_extrusion as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2}
                            group by anyo,  mes    
                            ORDER BY  anyo,  mes           
                                            """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'año', 'Mes', 'Unidades', 'Toneladas']
                cols_estadisticas = [2]
                cols_estadisticas2 = [3]

            if periodo == u'AÑO':
                sSql = """  SELECT date_part('year',c_ecd.fecha_inicio) as anyo,  sum(de.unidades_producto), trunc(sum(p.peso * de.unidades_producto / 1000000.0), 2)
                            FROM detalle_extrusion as de, producto as p, cabecera_proceso_ecd as c_ecd
                            WHERE p.id_producto = de.id_producto and c_ecd.uuid = de.uuid and de.activo = True
                                AND fecha_inicio >= '{0}' and fecha_inicio <= '{1}'  {2}
                            group by anyo    
                            ORDER BY  anyo         
                                            """.format(fecha_inicio, fecha_fin, cad_turno)

                cabeceras = [u'Año', 'Unidades', 'Toneladas']
                cols_estadisticas = [1]
                cols_estadisticas2 = [2]

        ManipularGrillas.setCantidadColumnasGrilla(self.grid_resultado_busqueda, len(cabeceras))
        ManipularGrillas.setCabecerasGrilla(self.grid_resultado_busqueda, cabeceras)
        rows = Ejecutar_SQL.select_varios_registros(sSql, 'frm_historial_procesos_ECD/buscar()', 500, BasesDeDatos.DB_PRINCIPAL)

        if rows == None:
            self.m_panel_resultados.Hide()
            return None, []

        self.rows_matriz = np.array(rows)
        if area_produccion == AREA_CARGUE_VAGONETAS:

            if periodo in ['DIA', 'AÑO']:
                # # Find indicies that you need to replace
                self.rows_matriz = np.where(self.rows_matriz == None, 0, self.rows_matriz)

                col_primera = self.rows_matriz[:, 1]
                col_rotura = self.rows_matriz[:, 2]

                col_total = col_primera + col_rotura

                col_ton_primera = self.rows_matriz[:, 3]
                col_ton_rotura = self.rows_matriz[:, 4]

                col_total_ton = col_ton_primera + col_ton_rotura

                col_porcentaje_primera = col_primera / col_total * 100
                col_porcentaje_rotura = col_rotura / col_total * 100

                self.rows_matriz = np.insert(self.rows_matriz, 1, col_total, axis = 1 )
                self.rows_matriz = np.insert(self.rows_matriz, 4, col_porcentaje_primera, axis=1)
                self.rows_matriz = np.insert(self.rows_matriz, 5, col_porcentaje_rotura, axis=1)
                self.rows_matriz = np.insert(self.rows_matriz, 6, col_total_ton, axis=1)

            if periodo in ['SEMANA', 'MES']:
                # # Find indicies that you need to replace
                self.rows_matriz = np.where(self.rows_matriz == None, 0, self.rows_matriz)

                col_primera = self.rows_matriz[:, 2]
                col_rotura = self.rows_matriz[:, 3]

                col_total = col_primera + col_rotura

                col_ton_primera = self.rows_matriz[:, 4]
                col_ton_rotura = self.rows_matriz[:, 5]

                col_total_ton = col_ton_primera + col_ton_rotura

                col_porcentaje_primera = col_primera / col_total * 100
                col_porcentaje_rotura = col_rotura / col_total * 100

                self.rows_matriz = np.insert(self.rows_matriz, 2, col_total, axis=1)
                self.rows_matriz = np.insert(self.rows_matriz, 5, col_porcentaje_primera, axis=1)
                self.rows_matriz = np.insert(self.rows_matriz, 6, col_porcentaje_rotura, axis=1)
                self.rows_matriz = np.insert(self.rows_matriz, 7, col_total_ton, axis=1)

        ##------------------------------------

        ManipularGrillas.setCantidadColumnasGrilla(self.grid_estadisticas, len(cabeceras_estadisticas))
        ManipularGrillas.setCabecerasGrilla(self.grid_estadisticas, cabeceras_estadisticas)

        ManipularGrillas.setCantidadColumnasGrilla(self.grid_estadisticas2, len(cabeceras_estadisticas2))
        ManipularGrillas.setCabecerasGrilla(self.grid_estadisticas2, cabeceras_estadisticas2)

        k=0
        for j in cols_estadisticas:
            columna = self.rows_matriz[:, j]

            sumatoria = round(np.sum(columna))
            self.grid_estadisticas.SetCellValue(0, k, str(sumatoria))

            media = round(np.mean(columna))
            self.grid_estadisticas.SetCellValue(1, k, str(media))

            desviacion_estandar = round(np.std(columna))
            self.grid_estadisticas.SetCellValue(2, k, str(desviacion_estandar))

            mediana = round(np.median(columna))
            self.grid_estadisticas.SetCellValue(3, k, str(mediana))

            maximo = round(np.max(columna))
            self.grid_estadisticas.SetCellValue(4, k, str(maximo))

            minimo = round(np.min(columna))
            self.grid_estadisticas.SetCellValue(5, k, str(minimo))

            n= len(columna)
            self.grid_estadisticas.SetCellValue(6, k, str(n))

            k += 1

        k = 0
        for j in cols_estadisticas2:
            columna = self.rows_matriz[:, j]

            sumatoria = round(np.sum(columna), 2)
            self.grid_estadisticas2.SetCellValue(0, k, str(sumatoria))

            media = round(np.mean(columna), 2)
            self.grid_estadisticas2.SetCellValue(1, k, str(media))

            desviacion_estandar = round(np.std(columna), 2)
            self.grid_estadisticas2.SetCellValue(2, k, str(desviacion_estandar))

            mediana = round(np.median(columna), 2)
            self.grid_estadisticas2.SetCellValue(3, k, str(mediana))

            maximo = round(np.max(columna), 2)
            self.grid_estadisticas2.SetCellValue(4, k, str(maximo))

            minimo = round(np.min(columna), 2)
            self.grid_estadisticas2.SetCellValue(5, k, str(minimo))

            n = len(columna)
            self.grid_estadisticas2.SetCellValue(6, k, str(n))

            k += 1

        ##------------------------------------
        rows = tuple(map(tuple, self.rows_matriz))

        return rows, cabeceras


    def cargar_checkList_turno(self, area_produccion):

        if area_produccion == 'DESPACHOS':
            self.checkList_turno.Clear()
        else:
            rows = DbGetVarios.listaTurnos(area_produccion, True)
            # id_turno, nom_turno, hora_inicio, hora_salida, activo

            if rows != None:
                la_lista = ManipularRows.crearListaValores(rows, 1)
                self.dic_turnos_todos = ManipularRows.crearDiccionarioTodosLosCampos(rows, 1)
                self.checkList_turno.Set(la_lista)


    def set_configuaracion_grid_resultado_busqueda(self):

        # list_columnas = [0, 1, 2, 3, 4, 5, 6]
        # ManipularGrillas.setColumnasSoloLectura(self.grid_resultado_busqueda, list_columnas)

        self.grid_resultado_busqueda.EnableEditing(False)

        self.grid_resultado_busqueda.AutoSizeColumns()


    def redondear_columna_grilla(self, list_cols=[], cant_decimales=0):
        cant_filas = self.grid_resultado_busqueda.GetNumberRows()
        
        for i in range(cant_filas):
            for j in list_cols:
                if cant_decimales == 0:
                    dato = round(float(self.grid_resultado_busqueda.GetCellValue(i, j)))
                else:
                    dato = round(float(self.grid_resultado_busqueda.GetCellValue(i, j)), cant_decimales)
                self.grid_resultado_busqueda.SetCellValue(i, j, str(dato))
        
        




